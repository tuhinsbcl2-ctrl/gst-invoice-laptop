"""
Excel/CSV export for invoices and reports.
"""
import os
import csv
import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def export_invoices_excel(invoices):
    """Export a list of Invoice objects to Excel bytes."""
    if not _HAS_OPENPYXL:
        return export_invoices_csv(invoices)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Invoices'

    headers = ['Invoice No', 'Date', 'Buyer Name', 'Place of Supply',
               'Taxable Value', 'CGST', 'SGST', 'IGST',
               'Round Off', 'Grand Total', 'Payment Status']
    bold = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold

    for row_idx, inv in enumerate(invoices, 2):
        ws.cell(row=row_idx, column=1, value=inv.invoice_no)
        ws.cell(row=row_idx, column=2, value=inv.date.strftime('%d-%b-%Y') if inv.date else '')
        ws.cell(row=row_idx, column=3, value=inv.buyer_name)
        ws.cell(row=row_idx, column=4, value=inv.place_of_supply)
        ws.cell(row=row_idx, column=5, value=inv.subtotal)
        ws.cell(row=row_idx, column=6, value=inv.cgst_total)
        ws.cell(row=row_idx, column=7, value=inv.sgst_total)
        ws.cell(row=row_idx, column=8, value=inv.igst_total)
        ws.cell(row=row_idx, column=9, value=inv.round_off)
        ws.cell(row=row_idx, column=10, value=inv.grand_total)
        ws.cell(row=row_idx, column=11, value=inv.payment_status)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), 'xlsx'


def export_invoices_csv(invoices):
    """Export a list of Invoice objects to CSV bytes."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Invoice No', 'Date', 'Buyer Name', 'Place of Supply',
                     'Taxable Value', 'CGST', 'SGST', 'IGST',
                     'Round Off', 'Grand Total', 'Payment Status'])
    for inv in invoices:
        writer.writerow([
            inv.invoice_no,
            inv.date.strftime('%d-%b-%Y') if inv.date else '',
            inv.buyer_name,
            inv.place_of_supply,
            inv.subtotal,
            inv.cgst_total,
            inv.sgst_total,
            inv.igst_total,
            inv.round_off,
            inv.grand_total,
            inv.payment_status,
        ])
    return output.getvalue().encode('utf-8'), 'csv'
